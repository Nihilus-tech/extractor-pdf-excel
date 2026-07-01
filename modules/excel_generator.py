# modules/excel_generator.py
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Colores del tema
COLOR_HEADER = "1E3A5F"      # Azul oscuro
COLOR_SUBHEADER = "2563EB"   # Azul medio
COLOR_FILA_PAR = "EFF6FF"    # Azul muy claro
BLANCO = "FFFFFF"
GRIS = "6B7280"

def _borde_delgado():
    lado = Side(style="thin", color="D1D5DB")
    return Border(left=lado, right=lado, top=lado, bottom=lado)

def _celda_header(ws, fila, col, texto, color_fondo=COLOR_HEADER):
    celda = ws.cell(row=fila, column=col, value=texto)
    celda.font = Font(bold=True, color=BLANCO, size=10)
    celda.fill = PatternFill("solid", fgColor=color_fondo)
    celda.alignment = Alignment(horizontal="center", vertical="center")
    celda.border = _borde_delgado()
    return celda

def generar_excel(datos):
    """
    Recibe el diccionario extraído por la IA y genera un Excel.
    Devuelve un BytesIO listo para que Flask lo envíe.
    """
    wb = openpyxl.Workbook()

    # ── HOJA 1: RESUMEN DEL DOCUMENTO ─────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Resumen"
    ws1.column_dimensions["A"].width = 25
    ws1.column_dimensions["B"].width = 40

    # Título
    ws1.merge_cells("A1:B1")
    titulo = ws1["A1"]
    titulo.value = f"DOCUMENTO: {datos.get('tipo_documento', 'DOCUMENTO').upper()}"
    titulo.font = Font(bold=True, color=BLANCO, size=13)
    titulo.fill = PatternFill("solid", fgColor=COLOR_HEADER)
    titulo.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 30

    # Campos del resumen
    campos = [
        ("Número de Documento", datos.get("numero_documento")),
        ("Fecha de Emisión", datos.get("fecha_emision")),
        ("Fecha de Vencimiento", datos.get("fecha_vencimiento")),
        ("Moneda", datos.get("moneda")),
        ("Método de Pago", datos.get("metodo_pago")),
        ("", ""),  # separador
        ("EMISOR", ""),
        ("Nombre Emisor", datos.get("emisor", {}).get("nombre")),
        ("RFC Emisor", datos.get("emisor", {}).get("rfc")),
        ("Dirección", datos.get("emisor", {}).get("direccion")),
        ("", ""),
        ("RECEPTOR", ""),
        ("Nombre Receptor", datos.get("receptor", {}).get("nombre")),
        ("RFC Receptor", datos.get("receptor", {}).get("rfc")),
        ("", ""),
        ("TOTALES", ""),
        ("Subtotal", datos.get("subtotal")),
        ("Impuestos", datos.get("impuestos")),
        ("TOTAL", datos.get("total")),
        ("", ""),
        ("Notas", datos.get("notas")),
    ]

    for i, (campo, valor) in enumerate(campos, start=2):
        fila = i
        if campo in ("EMISOR", "RECEPTOR", "TOTALES", ""):
            if campo:
                ws1.merge_cells(f"A{fila}:B{fila}")
                c = ws1[f"A{fila}"]
                c.value = campo
                c.font = Font(bold=True, color=BLANCO, size=9)
                c.fill = PatternFill("solid", fgColor=COLOR_SUBHEADER)
                c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            continue

        celda_campo = ws1.cell(row=fila, column=1, value=campo)
        celda_campo.font = Font(bold=True, size=9, color="1F2937")
        celda_campo.fill = PatternFill("solid", fgColor=COLOR_FILA_PAR)
        celda_campo.border = _borde_delgado()

        celda_valor = ws1.cell(row=fila, column=2, value=valor if valor is not None else "—")
        celda_valor.font = Font(size=9)
        celda_valor.border = _borde_delgado()

        # Formato de moneda para totales
        if campo in ("Subtotal", "Impuestos", "TOTAL") and isinstance(valor, (int, float)):
            celda_valor.number_format = '"$"#,##0.00'
            if campo == "TOTAL":
                celda_valor.font = Font(bold=True, size=10, color="1E3A5F")

    # ── HOJA 2: ITEMS / CONCEPTOS ─────────────────────────────────────────────
    ws2 = wb.create_sheet(title="Items")

    encabezados = ["#", "Descripción", "Cantidad", "Precio Unitario", "Importe"]
    anchos = [5, 45, 12, 18, 15]

    for col, (enc, ancho) in enumerate(zip(encabezados, anchos), start=1):
        ws2.column_dimensions[get_column_letter(col)].width = ancho
        _celda_header(ws2, 1, col, enc)

    ws2.row_dimensions[1].height = 22

    items = datos.get("items", [])
    if not items:
        ws2.cell(row=2, column=1, value="Sin items detectados").font = Font(color=GRIS, italic=True)
    else:
        for i, item in enumerate(items, start=1):
            fila = i + 1
            color_fondo = BLANCO if i % 2 == 0 else COLOR_FILA_PAR

            celdas_valores = [
                i,
                item.get("descripcion", ""),
                item.get("cantidad"),
                item.get("precio_unitario"),
                item.get("importe"),
            ]

            for col, valor in enumerate(celdas_valores, start=1):
                celda = ws2.cell(row=fila, column=col, value=valor)
                celda.fill = PatternFill("solid", fgColor=color_fondo)
                celda.border = _borde_delgado()
                celda.font = Font(size=9)

                # Formato monetario para columnas numéricas
                if col in (4, 5) and isinstance(valor, (int, float)):
                    celda.number_format = '"$"#,##0.00'
                    celda.alignment = Alignment(horizontal="right")
                elif col == 3:
                    celda.alignment = Alignment(horizontal="center")

        # Fila de total al final
        fila_total = len(items) + 2
        ws2.cell(row=fila_total, column=4, value="TOTAL:").font = Font(bold=True, size=10)
        celda_total = ws2.cell(row=fila_total, column=5, value=datos.get("total"))
        celda_total.font = Font(bold=True, size=10, color="1E3A5F")
        celda_total.number_format = '"$"#,##0.00'
        celda_total.fill = PatternFill("solid", fgColor=COLOR_FILA_PAR)
        celda_total.border = _borde_delgado()

    # Guardar en memoria
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer



def generar_excel_batch(resultados: list) -> BytesIO:
    """
    Recibe una lista de resultados del batch y genera un Excel
    con una hoja resumen + una hoja por documento.
    """
    wb = openpyxl.Workbook()

    # ── HOJA RESUMEN ──────────────────────────────────────────────────────────
    ws_resumen = wb.active
    ws_resumen.title = "Resumen General"

    # Anchos
    for col, ancho in zip("ABCDEFGH", [5, 35, 20, 20, 15, 15, 15, 20]):
        ws_resumen.column_dimensions[get_column_letter(ord(col) - 64)].width = ancho

    # Título
    ws_resumen.merge_cells("A1:H1")
    t = ws_resumen["A1"]
    t.value = f"RESUMEN DEL LOTE — {len(resultados)} DOCUMENTOS PROCESADOS"
    t.font = Font(bold=True, color=BLANCO, size=12)
    t.fill = PatternFill("solid", fgColor=COLOR_HEADER)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws_resumen.row_dimensions[1].height = 28

    # Encabezados de tabla resumen
    encabezados_resumen = ["#", "Archivo", "Tipo", "N° Documento", "Emisor", "Fecha", "Total", "Moneda"]
    for col, enc in enumerate(encabezados_resumen, start=1):
        _celda_header(ws_resumen, 2, col, enc, COLOR_SUBHEADER)

    total_global = 0.0

    for i, resultado in enumerate(resultados, start=1):
        d = resultado.get("datos", {})
        fila = i + 2
        color_fondo = BLANCO if i % 2 == 0 else COLOR_FILA_PAR

        total_doc = d.get("total")
        if isinstance(total_doc, (int, float)):
            total_global += total_doc

        valores = [
            i,
            resultado.get("filename", ""),
            d.get("tipo_documento", ""),
            d.get("numero_documento", ""),
            d.get("emisor", {}).get("nombre", "") if d.get("emisor") else "",
            d.get("fecha_emision", ""),
            total_doc,
            d.get("moneda", ""),
        ]

        for col, val in enumerate(valores, start=1):
            celda = ws_resumen.cell(row=fila, column=col, value=val)
            celda.fill = PatternFill("solid", fgColor=color_fondo)
            celda.border = _borde_delgado()
            celda.font = Font(size=9)
            if col == 7 and isinstance(val, (int, float)):
                celda.number_format = '"$"#,##0.00'
                celda.alignment = Alignment(horizontal="right")

    # Fila de total global
    fila_total = len(resultados) + 3
    ws_resumen.merge_cells(f"A{fila_total}:F{fila_total}")
    c_label = ws_resumen[f"A{fila_total}"]
    c_label.value = "TOTAL GLOBAL DEL LOTE"
    c_label.font = Font(bold=True, color=BLANCO, size=10)
    c_label.fill = PatternFill("solid", fgColor=COLOR_HEADER)
    c_label.alignment = Alignment(horizontal="right", vertical="center", indent=1)

    c_total = ws_resumen.cell(row=fila_total, column=7, value=total_global)
    c_total.font = Font(bold=True, color=BLANCO, size=11)
    c_total.fill = PatternFill("solid", fgColor=COLOR_HEADER)
    c_total.number_format = '"$"#,##0.00'
    c_total.alignment = Alignment(horizontal="right")
    c_total.border = _borde_delgado()

    # ── HOJA POR DOCUMENTO ────────────────────────────────────────────────────
    for resultado in resultados:
        d = resultado.get("datos", {})
        # Nombre de hoja: máx 31 chars, sin caracteres especiales
        nombre_hoja = resultado.get("filename", "doc")[:28].replace("/", "-").replace("\\", "-")
        ws = wb.create_sheet(title=nombre_hoja)

        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 38

        # Mini encabezado
        ws.merge_cells("A1:B1")
        h = ws["A1"]
        h.value = nombre_hoja
        h.font = Font(bold=True, color=BLANCO, size=10)
        h.fill = PatternFill("solid", fgColor=COLOR_SUBHEADER)
        h.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

        campos = [
            ("Tipo", d.get("tipo_documento")),
            ("N° Documento", d.get("numero_documento")),
            ("Fecha Emisión", d.get("fecha_emision")),
            ("Emisor", d.get("emisor", {}).get("nombre") if d.get("emisor") else None),
            ("RFC Emisor", d.get("emisor", {}).get("rfc") if d.get("emisor") else None),
            ("Receptor", d.get("receptor", {}).get("nombre") if d.get("receptor") else None),
            ("Subtotal", d.get("subtotal")),
            ("Impuestos", d.get("impuestos")),
            ("Total", d.get("total")),
            ("Moneda", d.get("moneda")),
            ("Método Pago", d.get("metodo_pago")),
        ]

        for i, (campo, valor) in enumerate(campos, start=2):
            c1 = ws.cell(row=i, column=1, value=campo)
            c1.font = Font(bold=True, size=9)
            c1.fill = PatternFill("solid", fgColor=COLOR_FILA_PAR)
            c1.border = _borde_delgado()

            c2 = ws.cell(row=i, column=2, value=valor if valor is not None else "—")
            c2.font = Font(size=9)
            c2.border = _borde_delgado()
            if campo in ("Subtotal", "Impuestos", "Total") and isinstance(valor, (int, float)):
                c2.number_format = '"$"#,##0.00'
                c2.alignment = Alignment(horizontal="right")

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer    